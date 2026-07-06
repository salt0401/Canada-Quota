"""Excel writer: create/update round-trips, the insert-shift formula bug
(known-issues Bug 1 + its incomplete first fix), the zero-country circular
reference, idempotency, and OVER flags. Everything asserts through
wb_invariants — the executable form of the known-issues checklist."""
import io
from datetime import date

from openpyxl import Workbook, load_workbook

import canada_trq_tracker as t
from wb_invariants import assert_trq_sheet_invariants, product_blocks

W1 = date(2026, 6, 29)
W2 = date(2026, 7, 6)
W3 = date(2026, 7, 13)


def _data(countries_by_product):
    """Build parse_trq_csv-shaped data for all 8 tracked products."""
    out = {}
    for i, prod in enumerate(t.TRACKED_PRODUCTS):
        cc = countries_by_product.get(prod, {})
        out[prod] = {"item_number": i + 1, "max_quota": 1_000_000 * (i + 1),
                     "max_share": 0.30, "total_util_pct": sum(cc.values()),
                     "countries": cc}
    return out


def _roundtrip(wb):
    """Save/load through a buffer so we assert what lands ON DISK."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


WEEK1 = _data({
    "Hot-Rolled Sheet": {"China": 0.01, "Taiwan": 0.05},
    "Steel Plate": {"China": 0.02},
    "Cold-Rolled Sheet": {"India": 0.03},
    "Coated Steel Sheet": {"Turkey": 0.04},
    "Rebar": {"Greece": 0.31},          # over the 30% cap -> OVER flag
    "Hot-Rolled Bar": {"Czechia": 0.01},
    "Structural Steel": {"Spain": 0.02},
    # Wire Rod: zero countries
})


def _fresh_sheet():
    wb = Workbook()
    del wb["Sheet"]
    t.create_trq_sheet(wb, "non-FTA Q1", "Q1", "NFTA", WEEK1, W1)
    return wb


def test_create_sheet_invariants_and_zero_country_total():
    wb = _roundtrip(_fresh_sheet())
    ws = wb["non-FTA Q1"]
    assert_trq_sheet_invariants(ws, t.FIRST_DATE_COL, t.FIRST_DATE_COL)
    blocks = product_blocks(ws)
    assert blocks["Wire Rod"]["countries"] == []
    v = ws.cell(row=blocks["Wire Rod"]["total_row"], column=t.FIRST_DATE_COL).value
    assert v == 0.0  # literal, not a formula


def _country_values(ws, col):
    """{(product, country): value} for one date column."""
    out = {}
    for prod, b in product_blocks(ws).items():
        for row, country in b["countries"]:
            out[(prod, country)] = ws.cell(row=row, column=col).value
    return out


def test_update_plain_week_two_columns():
    wb = _fresh_sheet()
    t.update_trq_sheet(wb["non-FTA Q1"], WEEK1, W2)
    ws = _roundtrip(wb)["non-FTA Q1"]
    assert_trq_sheet_invariants(ws, t.FIRST_DATE_COL, t.FIRST_DATE_COL + 1)
    assert ws.cell(row=2, column=t.FIRST_DATE_COL + 1).value == t.format_date_header(W2)
    assert ws.cell(row=2, column=t.FIRST_DATE_COL + 2).value == "OVER"
    # Historical week-1 values must survive the update untouched.
    week1_vals = _country_values(ws, t.FIRST_DATE_COL)
    assert week1_vals[("Hot-Rolled Sheet", "China")] == 0.01
    assert week1_vals[("Hot-Rolled Sheet", "Taiwan")] == 0.05
    assert week1_vals[("Rebar", "Greece")] == 0.31


def test_update_zero_country_product_no_circular_ref():
    """U2 regression: Wire Rod (TOTAL-only block) must get literal 0.0 in the
    new column, never =SUM(Fn:Fn) written into Fn itself."""
    wb = _fresh_sheet()
    t.update_trq_sheet(wb["non-FTA Q1"], WEEK1, W2)
    ws = _roundtrip(wb)["non-FTA Q1"]
    tr = product_blocks(ws)["Wire Rod"]["total_row"]
    for col in (t.FIRST_DATE_COL, t.FIRST_DATE_COL + 1):
        assert ws.cell(row=tr, column=col).value == 0.0


def test_new_countries_mid_sheet_do_not_corrupt_any_column():
    """U1 regression (known-issues Bug 1, complete fix): inserting new
    countries into EARLY products must leave every TOTAL formula in every
    date column spanning exactly its product's rows."""
    week2 = _data({
        "Hot-Rolled Sheet": {"Brazil": 0.02, "China": 0.015, "Taiwan": 0.06},  # +Brazil
        "Steel Plate": {"Austria": 0.01, "China": 0.025},                      # +Austria
        "Cold-Rolled Sheet": {"India": 0.035},
        "Coated Steel Sheet": {"Turkey": 0.045},
        "Rebar": {"Greece": 0.32},
        "Hot-Rolled Bar": {"Czechia": 0.012},
        "Structural Steel": {"Spain": 0.022},
    })
    wb = _fresh_sheet()
    t.update_trq_sheet(wb["non-FTA Q1"], week2, W2)
    ws = _roundtrip(wb)["non-FTA Q1"]

    assert_trq_sheet_invariants(ws, t.FIRST_DATE_COL, t.FIRST_DATE_COL + 1)
    blocks = product_blocks(ws)
    assert [c for _, c in blocks["Hot-Rolled Sheet"]["countries"]] == \
        ["China", "Taiwan", "Brazil"]  # appended above TOTAL
    # New country has no historical value in week-1 column
    brazil_row = blocks["Hot-Rolled Sheet"]["countries"][-1][0]
    assert ws.cell(row=brazil_row, column=t.FIRST_DATE_COL).value is None
    assert ws.cell(row=brazil_row, column=t.FIRST_DATE_COL + 1).value == 0.02
    # Every pre-existing week-1 value must still sit on its own country's row
    # after the inserts shifted rows (the value half of known-issues Bug 1).
    week1_vals = _country_values(ws, t.FIRST_DATE_COL)
    for prod, prod_data in WEEK1.items():
        for country, share in prod_data["countries"].items():
            assert week1_vals[(prod, country)] == share, (prod, country)


def test_zero_country_product_gains_first_country():
    wb = _fresh_sheet()
    week2 = dict(WEEK1)
    week2["Wire Rod"] = {"item_number": 7, "max_quota": 7_000_000,
                         "max_share": 0.30, "total_util_pct": 0.05,
                         "countries": {"Vietnam": 0.05}}
    t.update_trq_sheet(wb["non-FTA Q1"], week2, W2)
    ws = _roundtrip(wb)["non-FTA Q1"]
    assert_trq_sheet_invariants(ws, t.FIRST_DATE_COL, t.FIRST_DATE_COL + 1)
    blocks = product_blocks(ws)
    assert [c for _, c in blocks["Wire Rod"]["countries"]] == ["Vietnam"]


def test_three_weeks_with_inserts_all_columns_clean():
    wb = _fresh_sheet()
    week2 = _data({**{p: WEEK1[p]["countries"] for p in t.TRACKED_PRODUCTS},
                   "Hot-Rolled Sheet": {"Brazil": 0.02, "China": 0.015, "Taiwan": 0.06}})
    t.update_trq_sheet(wb["non-FTA Q1"], week2, W2)
    week3 = _data({**{p: week2[p]["countries"] for p in t.TRACKED_PRODUCTS},
                   "Steel Plate": {"Austria": 0.01, "China": 0.03}})
    t.update_trq_sheet(wb["non-FTA Q1"], week3, W3)
    ws = _roundtrip(wb)["non-FTA Q1"]
    assert_trq_sheet_invariants(ws, t.FIRST_DATE_COL, t.FIRST_DATE_COL + 2)


def _sheet_snapshot(ws):
    return [[(c.value, c.number_format) for c in row] for row in ws.iter_rows()]


def test_same_day_update_is_idempotent():
    """The same-date guard's contract is NO mutation at all on re-run, not
    just an unchanged column count."""
    wb = _fresh_sheet()
    t.update_trq_sheet(wb["non-FTA Q1"], WEEK1, W2)
    before = _sheet_snapshot(wb["non-FTA Q1"])
    t.update_trq_sheet(wb["non-FTA Q1"], WEEK1, W2)
    assert _sheet_snapshot(wb["non-FTA Q1"]) == before


def test_update_heals_stale_formulas_without_inserts():
    """_rewrite_total_formulas must repair pre-existing corruption (e.g. left
    by the pre-2026-07-06 code) even on a week with NO new countries."""
    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    blocks = product_blocks(ws)
    tr = blocks["Steel Plate"]["total_row"]
    # simulate legacy insert-shift corruption in the historical column
    ws.cell(row=tr, column=t.FIRST_DATE_COL).value = "=SUM(E3:E3)"
    t.update_trq_sheet(ws, WEEK1, W2)   # no new countries this week
    healed = _roundtrip(wb)["non-FTA Q1"]
    assert_trq_sheet_invariants(healed, t.FIRST_DATE_COL, t.FIRST_DATE_COL + 1)


def test_over_flags():
    ws = _roundtrip(_fresh_sheet())["non-FTA Q1"]
    blocks = product_blocks(ws)
    over_col = t.FIRST_DATE_COL + 1
    greece_row = blocks["Rebar"]["countries"][0][0]
    assert ws.cell(row=greece_row, column=over_col).value == "YES"   # 0.31 >= 0.30
    china_row = blocks["Steel Plate"]["countries"][0][0]
    assert ws.cell(row=china_row, column=over_col).value is None     # 0.02 < 0.30
    assert ws.cell(row=blocks["Rebar"]["total_row"], column=over_col).value is None


def test_write_text_blocks_formula_injection():
    wb = Workbook()
    ws = wb.active
    t._write_text(ws, 1, 1, "=HYPERLINK(evil)")
    t._write_text(ws, 2, 1, "Türkiye")
    ws2 = _roundtrip(wb).active
    assert ws2.cell(row=1, column=1).value == "=HYPERLINK(evil)"
    assert ws2.cell(row=1, column=1).data_type == "s"  # stored as string, not formula
    assert ws2.cell(row=2, column=1).value == "Türkiye"


def test_b1_sheet_year_from_quarter_end_date():
    """The B1 window year always equals the TRQ quarter's end-date year —
    correct even for a Q3 run in December (upcoming window is next year) and
    for a stale landing page leaving Q4 selected in July (window that closed
    THIS year, where a today-based heuristic would say next year)."""
    wb = Workbook()
    cases = [
        ("Q3", date(2026, 12, 28), date(2027, 3, 25), "March 31, 2027"),
        ("Q3", date(2027, 1, 4), date(2027, 3, 25), "March 31, 2027"),
        ("Q1", date(2026, 7, 6), date(2026, 9, 29), "September 30, 2026"),
        ("Q4", date(2026, 7, 2), date(2026, 6, 27), "June 30, 2026"),  # stale page
    ]
    for quarter, today, q_end, expect in cases:
        t.create_b1_sheet(wb, None, quarter, today, quarter_end=q_end)
        assert expect in wb["B1 Imports"].cell(row=1, column=1).value, (quarter, today)


def test_update_missing_over_column_fails_loud():
    """A hand-edited sheet without its OVER header must raise, not silently
    skip this and every future weekly update."""
    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    ws.cell(row=2, column=t.FIRST_DATE_COL + 1).value = None  # delete OVER header
    import pytest
    with pytest.raises(RuntimeError, match="OVER"):
        t.update_trq_sheet(ws, WEEK1, W2)


def test_date_headers_not_yellow_after_update():
    """Only the OVER header is highlighted; a new date header written into the
    old OVER cell must not inherit its yellow fill."""
    wb = _fresh_sheet()
    t.update_trq_sheet(wb["non-FTA Q1"], WEEK1, W2)
    ws = _roundtrip(wb)["non-FTA Q1"]
    assert ws.cell(row=2, column=t.FIRST_DATE_COL + 1).fill.fill_type is None
    over_cell = ws.cell(row=2, column=t.FIRST_DATE_COL + 2)
    assert over_cell.value == "OVER" and over_cell.fill.fill_type == "solid"


# --- _load_or_init_workbook guard --------------------------------------------

def test_missing_workbook_fails_loud(tmp_path, monkeypatch):
    import pytest
    monkeypatch.setattr(t, "OUTPUT_FILE", str(tmp_path / "wb.xlsx"))
    monkeypatch.delenv("TRQ_ALLOW_NEW_WORKBOOK", raising=False)
    with pytest.raises(SystemExit) as e:
        t._load_or_init_workbook()
    assert e.value.code == 1   # non-zero is the contract: the commit step must not run
    monkeypatch.setenv("TRQ_ALLOW_NEW_WORKBOOK", "1")
    wb = t._load_or_init_workbook()   # deliberate init is allowed
    assert wb.sheetnames == []


def test_existing_workbook_loads_and_corrupt_fails(tmp_path, monkeypatch):
    import pytest
    path = tmp_path / "wb.xlsx"
    monkeypatch.setattr(t, "OUTPUT_FILE", str(path))
    _fresh_sheet().save(path)
    assert "non-FTA Q1" in t._load_or_init_workbook().sheetnames
    path.write_bytes(b"not a zip archive")
    with pytest.raises(SystemExit) as e:
        t._load_or_init_workbook()
    assert e.value.code == 1


# --- validate_workbook (the post-save gate) -----------------------------------

def _saved(tmp_path, wb):
    p = str(tmp_path / "gate.xlsx")
    wb.save(p)
    return p


def test_gate_passes_good_workbook(tmp_path):
    t.validate_workbook(_saved(tmp_path, _fresh_sheet()))


def test_gate_rejects_each_corruption(tmp_path):
    import pytest
    col = t.FIRST_DATE_COL

    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    tr = product_blocks(ws)["Steel Plate"]["total_row"]
    ws.cell(row=tr, column=col).value = "=SUM(E3:E3)"      # stale/shifted range
    with pytest.raises(RuntimeError, match="stale/shifted"):
        t.validate_workbook(_saved(tmp_path, wb))

    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    china_row = product_blocks(ws)["Hot-Rolled Sheet"]["countries"][0][0]
    ws.cell(row=china_row, column=col).value = "=SUM(A1:A2)"   # formula in data cell
    with pytest.raises(RuntimeError, match="numeric or blank"):
        t.validate_workbook(_saved(tmp_path, wb))

    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    wr = product_blocks(ws)["Wire Rod"]["total_row"]
    ws.cell(row=wr, column=col).value = f"=SUM(E{wr}:E{wr})"   # circular self-sum
    with pytest.raises(RuntimeError, match="not 0"):
        t.validate_workbook(_saved(tmp_path, wb))

    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    ws.cell(row=2, column=t.FIRST_DATE_COL + 1).value = None   # no OVER header
    with pytest.raises(RuntimeError, match="no OVER header"):
        t.validate_workbook(_saved(tmp_path, wb))

    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    china_row = product_blocks(ws)["Hot-Rolled Sheet"]["countries"][0][0]
    ws.insert_rows(china_row + 1)                               # duplicate country
    ws.cell(row=china_row + 1, column=1).value = "Hot-Rolled Sheet"
    ws.cell(row=china_row + 1, column=2).value = "China"
    with pytest.raises(RuntimeError, match="duplicate"):
        t.validate_workbook(_saved(tmp_path, wb))


def test_gate_checks_historical_columns_too(tmp_path):
    """The gate's whole point is EVERY date column — a newest-only check is
    exactly the blind spot that hid the original Bug 1 corruption."""
    import pytest
    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    t.update_trq_sheet(ws, WEEK1, W2)   # two date columns now
    tr = product_blocks(ws)["Steel Plate"]["total_row"]
    ws.cell(row=tr, column=t.FIRST_DATE_COL).value = "=SUM(E3:E3)"  # corrupt OLD column only
    with pytest.raises(RuntimeError, match="stale/shifted"):
        t.validate_workbook(_saved(tmp_path, wb))


def test_gate_rejects_country_row_without_product_label(tmp_path):
    """A hand-edited row (country in B, blank A) is counted by the writer's
    range logic but invisible to the gate's walker — it must fail with an
    actionable message, not a misleading stale-formula error."""
    import pytest
    wb = _fresh_sheet()
    ws = wb["non-FTA Q1"]
    china_row = product_blocks(ws)["Hot-Rolled Sheet"]["countries"][0][0]
    ws.insert_rows(china_row + 1)
    ws.cell(row=china_row + 1, column=2).value = "Chile"   # column A left blank
    ws.cell(row=china_row + 1, column=t.FIRST_DATE_COL).value = 0.02
    with pytest.raises(RuntimeError, match="no product"):
        t.validate_workbook(_saved(tmp_path, wb))


def test_gate_ignores_non_trq_sheets(tmp_path):
    wb = _fresh_sheet()
    t.create_b1_sheet(wb, None, "Q1", W1, quarter_end=date(2026, 9, 29))
    t.create_hts_sheet(wb)
    t.validate_workbook(_saved(tmp_path, wb))


def test_b1_sheet_year_heuristic_fallback():
    """Without a quarter_end (legacy call shape), the today-based heuristic
    still handles the December Q3 case."""
    wb = Workbook()
    t.create_b1_sheet(wb, None, "Q3", date(2026, 12, 28))
    assert "March 31, 2027" in wb["B1 Imports"].cell(row=1, column=1).value
    t.create_b1_sheet(wb, None, "Q1", date(2026, 7, 6))
    assert "September 30, 2026" in wb["B1 Imports"].cell(row=1, column=1).value
