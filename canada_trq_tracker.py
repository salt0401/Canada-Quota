"""
Canada TRQ Weekly Tracker

Downloads Canadian steel Tariff Rate Quota (TRQ) utilization data and B1 import
data, parsing them into an Excel workbook that matches Laura's template format.
Designed to run weekly via GitHub Actions.
"""

import csv
import io
import logging
import os
import platform
import re
import sys
import time
from datetime import date
from urllib.parse import urljoin, urlparse

import requests
from requests.adapters import HTTPAdapter

try:
    from urllib3.util.retry import Retry
except ImportError:  # urllib3 vendored inside requests on some installs
    from requests.packages.urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

BASE_URL = "https://www.eics-scei.gc.ca/report-rapport"
B1_URL = f"{BASE_URL}/b1.htm"
# Official landing page that lists the current report for each quarter with its
# real filename and date range. Discovery reads from here so the scraper follows
# the source's year-over-year file renames (e.g. Q1 -> Y2Q1) instead of guessing.
LANDING_URL = (
    "https://www.international.gc.ca/trade-commerce/controls-controles/"
    "steel-acier/trq_data-donnees_tarifaires.aspx?lang=eng"
)
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "canada_trq_tracker.xlsx")

# ---------------------------------------------------------------------------
# HTTP session
# ---------------------------------------------------------------------------
# The source host (www.eics-scei.gc.ca) sits behind a firewall that
# intermittently drops connections from cloud/datacenter egress IPs (e.g. some
# GitHub Actions runners), which surfaces as a TCP *connect timeout* rather than
# an HTTP error. A browser-like User-Agent plus bounded retries with exponential
# backoff smooth over transient drops and slow responses. NOTE: a hard block of
# the runner's IP cannot be fixed here — it requires a fresh egress IP, i.e. a
# re-run of the job on a new runner. See docs/known-issues.md.
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/csv,text/html,application/xhtml+xml,*/*;q=0.8",
}
# (connect, read) timeouts: fail a blocked connect quickly, allow slow reads.
CSV_TIMEOUT = (10, 60)
HTML_TIMEOUT = (10, 120)  # landing page and B1 are large/slow HTML pages


def _make_session() -> requests.Session:
    """Build a requests session with browser UA and retry/backoff on both
    connection failures and retryable HTTP statuses."""
    retry = Retry(
        total=3,
        connect=2,
        read=2,
        status=2,
        backoff_factor=1.5,  # sleeps ~1.5s, 3s, 6s between attempts
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET"]),
        raise_on_status=False,
    )
    session = requests.Session()
    session.headers.update(REQUEST_HEADERS)
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


SESSION = _make_session()

# The 8 product categories Laura tracks (must match TRQ product names exactly)
TRACKED_PRODUCTS = [
    "Hot-Rolled Sheet",
    "Steel Plate",
    "Cold-Rolled Sheet",
    "Coated Steel Sheet",
    "Rebar",
    "Hot-Rolled Bar",
    "Wire Rod",
    "Structural Steel",
]

# HTS code mapping (dotted format from Laura's template)
HTS_CODE_MAP = {
    "Hot-Rolled Sheet": [
        "7208.10.00.00", "7208.25.00.00", "7208.26.00.00", "7208.27.00.00",
        "7208.36.00.00", "7208.37.00.00", "7208.38.00.00", "7208.39.00.00",
    ],
    "Steel Plate": [
        "7208.40.00.00", "7208.51.00.00", "7208.52.00.00",
    ],
    "Cold-Rolled Sheet": [
        "7209.15.00.00", "7209.16.00.00", "7209.17.00.00", "7209.18.00.00",
        "7209.25.00.00", "7209.26.00.00", "7209.27.00.00", "7209.28.00.00",
    ],
    "Coated Steel Sheet": [
        "7210.41.00.00", "7210.49.00.00", "7210.61.00.00",
        "7210.69.00.00", "7210.70.00.00", "7210.90.00.00",
    ],
    "Rebar": ["7214.20.00.00"],
    "Hot-Rolled Bar": [
        "7214.30.00.00", "7214.91.00.00", "7214.99.00.00",
    ],
    "Wire Rod": [
        "7213.10.00.00", "7213.20.00.00", "7213.91.00.00", "7213.99.00.00",
    ],
    "Structural Steel": [
        "7216.10.00.00", "7216.21.00.00", "7216.22.00.00", "7216.31.00.00",
        "7216.32.00.00", "7216.33.00.00", "7216.40.00.00", "7216.50.00.00",
        "7216.69.00.00",
    ],
}

# Build reverse lookup: 8-digit HS prefix -> product category
# Laura's codes use "00" as the last 2 statistical digits (e.g., 7214.20.00.00 → 7214200000)
# but B1 uses specific statistical suffixes (e.g., 7214200011, 7214200012).
# Matching on first 8 digits captures all sub-codes correctly.
_HTS_REVERSE: dict[str, str] = {}
for _prod, _codes in HTS_CODE_MAP.items():
    for _code in _codes:
        _prefix_8 = _code.replace(".", "")[:8]
        _HTS_REVERSE[_prefix_8] = _prod

# TRQ offset quarter -> overlapping B1 calendar quarter months
TRQ_TO_B1_CALENDAR_MONTHS: dict[str, tuple[int, int]] = {
    "Q1": (7, 9),
    "Q2": (10, 12),
    "Q3": (1, 3),
    "Q4": (4, 6),
}

# Country name normalization (B1 name -> TRQ name)
COUNTRY_NAME_MAP: dict[str, str] = {
    # Both sources use the same government system; add entries if mismatches found.
}

# Sheet-name prefix per TRQ partner group. main() uses this both to name the
# sheets and to detect a quarter transition, so the two must never diverge.
TRQ_SHEET_PREFIX = {"NFTA": "non-FTA", "FTA": "FTA"}

# Excel formatting constants
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
BOLD_FONT = Font(bold=True)

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def parse_number(s: str) -> float:
    """Strip commas and parse a numeric string. Returns 0.0 for empty/blank.
    Handles 'Max Utilized' as 0.0 (actual KGM value unknown)."""
    s = s.strip().strip('"')
    if not s or s == "Max Utilized":
        return 0.0
    return float(s.replace(",", ""))


def parse_percent(s: str) -> float:
    """Parse a percentage string to a decimal (e.g. '2.95%' -> 0.0295).
    Handles 'Max Utilized' as 1.0 (100%)."""
    s = s.strip().strip('"')
    if not s:
        return 0.0
    if s == "Max Utilized":
        return 1.0
    if s.endswith("%"):
        return float(s[:-1].replace(",", "")) / 100.0
    # Already a decimal number
    return float(s.replace(",", ""))


def format_date_header(d: date) -> str:
    """Format a date as 'March 30 2026' (full month, no leading zero on day)."""
    if platform.system() == "Windows":
        return d.strftime("%B %#d %Y")
    return d.strftime("%B %-d %Y")


# NOTE: the calendar-based get_current_quarter() was removed. The TRQ "program
# year" rolls over and the source renames quarter files (Q1 -> Y2Q1) and shifts
# boundary dates each year, so deriving the quarter from the calendar silently
# tracked the wrong/closed quarter after a rollover. The current quarter, its
# real CSV URL, and its date label are now discovered from the landing page; see
# discover_current_reports().


def get_quarter_date_range(quarter: str, ref_date: date) -> tuple[str, str]:
    """Stale fallback for human-readable quarter dates (real boundaries drift
    year to year). Discovery supplies the authoritative date strings; this is
    only a default for create_trq_sheet when none are passed."""
    year = ref_date.year
    if quarter == "Q1":
        return f"June 27, {year}", f"September 25, {year}"
    if quarter == "Q2":
        return f"September 26, {year}", f"December 25, {year}"
    if quarter == "Q3":
        # Q3 spans the year boundary
        start_year = year - 1 if ref_date.month <= 3 else year
        end_year = start_year + 1 if start_year == year else year
        return f"December 26, {start_year}", f"March 25, {end_year}"
    # Q4
    return f"March 26, {year}", f"June 26, {year}"


def should_fetch_b1(trq_quarter: str, today: date) -> bool:
    """Check if B1 page currently shows data for the calendar quarter matching this TRQ quarter."""
    cal_start, cal_end = TRQ_TO_B1_CALENDAR_MONTHS[trq_quarter]
    return cal_start <= today.month <= cal_end


# ---------------------------------------------------------------------------
# TRQ CSV Download & Parsing
# ---------------------------------------------------------------------------

def download_csv(url: str) -> str | None:
    """Download a TRQ CSV by full URL. Returns text, or None on HTTP 404 or
    any download failure that survives the session's retries."""
    try:
        log.info("Downloading %s ...", url)
        resp = SESSION.get(url, timeout=CSV_TIMEOUT)
        if resp.status_code == 404:
            log.warning("HTTP 404 for %s", url)
            return None
        resp.raise_for_status()
        resp.encoding = "utf-8"
        return resp.text
    except requests.RequestException as exc:
        log.warning("Download failed for %s: %s", url, exc)
        return None


# Index 0 is a placeholder so list index == month number.
_MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
_MONTHS = {name: num for num, name in enumerate(_MONTH_NAMES[1:], start=1)}


def _parse_quarter_label(text: str) -> tuple[str, date, date, str, str] | None:
    """Parse a landing-page label such as
        'Quarter 1: June 28 – September 29, 2026'
        'Quarter 3: December 26, 2025 – March 25, 2026'
    into (quarter_label, start_date, end_date, start_str, end_str), or None.
    Handles the en-dash separator, a missing start-year (inferred from the end),
    and year-wrap quarters. Only 'Quarter N' rows match, so sub-period
    'Period N' rows are ignored."""
    m = re.search(r"Quarter\s+(\d)\s*:\s*(.+)", text)
    if not m:
        return None
    qnum = m.group(1)
    parts = re.split(r"\s*[–—-]\s*", m.group(2).strip(), maxsplit=1)
    if len(parts) != 2:
        return None
    end_m = re.match(r"([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})", parts[1].strip())
    start_m = re.match(r"([A-Za-z]+)\s+(\d{1,2})(?:,\s*(\d{4}))?", parts[0].strip())
    if not end_m or not start_m:
        return None
    e_mon, e_day, e_year = end_m.group(1), int(end_m.group(2)), int(end_m.group(3))
    s_mon, s_day = start_m.group(1), int(start_m.group(2))
    if s_mon not in _MONTHS or e_mon not in _MONTHS:
        return None
    s_year = int(start_m.group(3)) if start_m.group(3) else (
        e_year - 1 if _MONTHS[s_mon] > _MONTHS[e_mon] else e_year)
    try:
        start_date = date(s_year, _MONTHS[s_mon], s_day)
        end_date = date(e_year, _MONTHS[e_mon], e_day)
    except ValueError:
        return None
    return (f"Q{qnum}", start_date, end_date,
            f"{s_mon} {s_day}, {s_year}", f"{e_mon} {e_day}, {e_year}")


def discover_current_reports(today: date | None = None) -> dict:
    """Discover the CURRENT-quarter TRQ CSV for both partner groups from the
    official landing page. Filenames/labels/dates are READ from the page, never
    constructed -- the source renames files across program years (e.g. Q1 ->
    Y2Q1) and uses inconsistent casing.

    Returns {"FTA": {...}, "NFTA": {...}} where each value has keys csv_url,
    quarter, start_date, end_date, start_str, end_str.

    Raises RuntimeError on ANY problem (page unreachable, structure changed, no
    current report, malformed CSV link, or FTA/NFTA disagreeing on the quarter)
    so the caller can FAIL LOUD rather than silently publish a stale quarter."""
    if today is None:
        today = date.today()
    try:
        resp = SESSION.get(LANDING_URL, timeout=HTML_TIMEOUT)
        resp.raise_for_status()
        resp.encoding = "utf-8"
    except requests.RequestException as exc:
        raise RuntimeError(f"could not fetch landing page {LANDING_URL}: {exc}") from exc

    soup = BeautifulSoup(resp.text, "lxml")

    # Group anchors by report stem (e.g. TRQ_FTA-Y2Q1). Each quarter row has a
    # "Quarter N: <dates>" link (-> .htm) plus an adjacent "[.CSV]" link; pair
    # them by shared stem. The CSV link is identified by its "[.CSV]" text, NOT
    # by extension, because the page has at least one malformed [.CSV] -> .htm.
    rows: dict[str, dict] = {}
    for a in soup.find_all("a", href=True):
        # Resolve relative hrefs against the landing page so the returned
        # csv_url is always absolute and fetchable.
        href = urljoin(LANDING_URL, a["href"].strip())
        base = href.rsplit("/", 1)[-1]
        if not base.upper().startswith("TRQ_"):
            continue
        stem = re.sub(r"\.(html?|csv|xlsx)$", "", base, flags=re.I).upper()
        row = rows.setdefault(stem, {"label": None, "csv_url": None})
        text = a.get_text(" ", strip=True)
        if re.search(r"Quarter\s+\d", text):
            row["label"] = text
        elif text.replace(" ", "").upper() == "[.CSV]":
            row["csv_url"] = href

    candidates: dict[str, list] = {"FTA": [], "NFTA": []}
    for stem, row in rows.items():
        if not row["label"] or not row["csv_url"]:
            continue
        parsed = _parse_quarter_label(row["label"])
        if not parsed:
            continue
        quarter, start_date, end_date, start_str, end_str = parsed
        partner = "NFTA" if "NFTA" in stem else ("FTA" if "FTA" in stem else None)
        if partner is None:
            continue
        candidates[partner].append({
            "csv_url": row["csv_url"], "quarter": quarter,
            "start_date": start_date, "end_date": end_date,
            "start_str": start_str, "end_str": end_str,
        })

    result = {}
    for partner in ("FTA", "NFTA"):
        cands = candidates[partner]
        if not cands:
            raise RuntimeError(f"no {partner} 'Quarter N' reports found on landing page")
        # Current = the report whose date range contains today; if today falls in
        # a gap between quarters, take the most recently started quarter.
        in_range = [c for c in cands if c["start_date"] <= today <= c["end_date"]]
        pool = in_range or [c for c in cands if c["start_date"] <= today]
        if not pool:
            raise RuntimeError(f"no current {partner} TRQ quarter found for {today}")
        chosen = max(pool, key=lambda c: c["start_date"])
        if not chosen["csv_url"].lower().endswith(".csv"):
            raise RuntimeError(
                f"{partner} current report CSV link is not a .csv (malformed page "
                f"link): {chosen['csv_url']}")
        # Only ever download from the government's own domain; a page compromise
        # or scraper bug must not make us fetch (and publish) foreign data.
        host = urlparse(chosen["csv_url"]).hostname or ""
        if host != "gc.ca" and not host.endswith(".gc.ca"):
            raise RuntimeError(
                f"{partner} current report CSV link points off the gc.ca domain: "
                f"{chosen['csv_url']}")
        result[partner] = chosen

    if result["FTA"]["quarter"] != result["NFTA"]["quarter"]:
        raise RuntimeError(
            f"FTA/NFTA resolved to different quarters "
            f"({result['FTA']['quarter']} vs {result['NFTA']['quarter']}); aborting")
    return result


# Part A (the 23-product summary table) starts on CSV line 5 in both formats.
_PART_A_START = 4  # 0-indexed
# The TRQ program publishes exactly this many product rows. A different count
# means the source structure changed, which voids the POSITIONAL Part A/Part B
# match below — publishing anyway would assign countries to the wrong products.
_EXPECTED_PART_A_ITEMS = 23


def parse_trq_csv(csv_text: str) -> dict:
    """Parse TRQ CSV (either format) and return structured data.

    Returns:
        {
            "Hot-Rolled Sheet": {
                "item_number": 3,
                "max_quota": 2370500,
                "max_share": 0.41,
                "total_util_pct": 0.4397,
                "countries": {"China": 0.0295, "India": 0.0011, ...}
            },
            ...
        }
    """
    lines = csv_text.splitlines()
    if not lines:
        raise ValueError("Empty CSV")

    # Detect format
    if lines[0].startswith("ExecutionTime"):
        fmt = "old"
        part_a_offset = 0
        part_b_offset = 0
    elif lines[0].startswith("Textbox"):
        fmt = "new"
        part_a_offset = 6
        part_b_offset = 3
    else:
        raise ValueError(f"Unrecognized CSV format. Line 1: {lines[0][:80]}")

    log.info("CSV format detected: %s", fmt)

    # --- Part A: from line 5 to the first blank line ---
    # The boundary is located dynamically (not hard-coded) so a source-side row
    # add/remove shifts the Part B walker with it instead of desynchronizing;
    # the count assert below then fails the run loudly.
    part_a_items: list[dict] = []
    i = _PART_A_START
    while i < len(lines) and lines[i].strip():
        fields = list(csv.reader(io.StringIO(lines[i])))[0]
        o = part_a_offset
        if len(fields) < o + 7:
            log.warning("Part A line %d has only %d fields (expected %d+), skipping: %s",
                        i + 1, len(fields), o + 7, lines[i][:80])
            i += 1
            continue
        part_a_items.append({
            "item_number": int(fields[o]),
            "product_name": fields[o + 1].strip(),
            "max_quota": int(parse_number(fields[o + 2])),
            "max_share": parse_percent(fields[o + 3]),
            # parse_number/parse_percent already map "Max Utilized" to 0.0 / 1.0
            "total_util_kgm": parse_number(fields[o + 4]),
            "total_util_pct": parse_percent(fields[o + 5]),
        })
        i += 1

    if len(part_a_items) != _EXPECTED_PART_A_ITEMS:
        raise ValueError(
            f"Part A has {len(part_a_items)} product rows (expected "
            f"{_EXPECTED_PART_A_ITEMS}); the CSV structure changed and the "
            f"positional Part A/Part B match cannot be trusted")

    # --- Part B: everything after the blank line that ends Part A ---
    part_b_start = i + 1
    # Walk Part B sections, matching sequentially to Part A items
    part_b_sections: list[list[dict]] = []
    current_section: list[dict] | None = None

    for i in range(part_b_start, len(lines)):
        line = lines[i].strip()

        if not line:
            # Blank line: end of current section
            if current_section is not None:
                part_b_sections.append(current_section)
                current_section = None
            continue

        # Check if this is a metadata header line
        if line.startswith("Textbox") or line.startswith("CONTROL_ITEMS_Level_1"):
            if current_section is not None:
                # Previous section ended (no blank line separator — shouldn't happen, but be safe)
                part_b_sections.append(current_section)
            current_section = []
            continue

        # Data row (not blank, not header)
        if current_section is not None:
            fields = list(csv.reader(io.StringIO(line)))[0]
            o = part_b_offset
            if len(fields) > o + 4:
                country = fields[o].strip()
                util_kgm = parse_number(fields[o + 1])
                share_pct = parse_percent(fields[o + 2])
                total_kgm = parse_number(fields[o + 3])
                total_pct = parse_percent(fields[o + 4])
                current_section.append({
                    "country": country,
                    "util_kgm": util_kgm,
                    "share_pct": share_pct,
                    "total_kgm": total_kgm,
                    "total_pct": total_pct,
                })
            else:
                # Symmetric with the Part A guard: never drop a row silently.
                log.warning("Part B line %d has only %d fields (expected %d+), skipping: %s",
                            i + 1, len(fields), o + 5, line[:80])

    # Don't forget the last section if file doesn't end with blank line
    if current_section is not None:
        part_b_sections.append(current_section)

    # --- Match Part B sections to Part A items ---
    result: dict[str, dict] = {}
    b_idx = 0

    for a_item in part_a_items:
        product_name = a_item["product_name"]
        countries: dict[str, float] = {}

        if a_item["total_util_kgm"] > 0 or a_item["total_util_pct"] > 0:
            # This product should have a Part B section
            if b_idx < len(part_b_sections):
                section = part_b_sections[b_idx]
                b_idx += 1

                for row in section:
                    countries[row["country"]] = row["share_pct"]

                # Validation: compare total_kgm (skip if "Max Utilized" made Part A KGM = 0)
                if section and a_item["total_util_kgm"] > 0:
                    b_total_kgm = section[0]["total_kgm"]
                    if abs(b_total_kgm - a_item["total_util_kgm"]) > 1:
                        log.warning(
                            "Part A/B mismatch for %s: Part A util=%.0f, Part B total=%.0f",
                            product_name, a_item["total_util_kgm"], b_total_kgm,
                        )
            else:
                # Sections exhausted: nothing to consume. Alignment for any
                # later products is already lost; the warning is the signal.
                log.warning("No Part B section found for %s (expected non-zero utilization)", product_name)
        else:
            # Zero utilization — still has a header in Part B but no data rows.
            # The header+blank was counted as an empty section.
            if b_idx < len(part_b_sections) and len(part_b_sections[b_idx]) == 0:
                b_idx += 1  # consume the empty section

        # Only store tracked products
        if product_name in TRACKED_PRODUCTS:
            result[product_name] = {
                "item_number": a_item["item_number"],
                "max_quota": a_item["max_quota"],
                "max_share": a_item["max_share"],
                "total_util_pct": a_item["total_util_pct"],
                "countries": countries,
            }

    # Validate: all tracked products should be present
    for prod in TRACKED_PRODUCTS:
        if prod not in result:
            log.warning("Tracked product '%s' not found in CSV data", prod)
            result[prod] = {
                "item_number": 0,
                "max_quota": 0,
                "max_share": 0.0,
                "total_util_pct": 0.0,
                "countries": {},
            }

    # Data validation
    for prod_name, prod_data in result.items():
        for country, share in prod_data["countries"].items():
            if not (0.0 <= share <= 1.0):
                log.warning("Share out of range for %s/%s: %.4f", prod_name, country, share)
        country_sum = sum(prod_data["countries"].values())
        if abs(country_sum - prod_data["total_util_pct"]) > 0.01:
            log.warning(
                "Country share sum (%.4f) != total utilization (%.4f) for %s",
                country_sum, prod_data["total_util_pct"], prod_name,
            )

    return result


# ---------------------------------------------------------------------------
# B1 Import Data Scraping
# ---------------------------------------------------------------------------

def scrape_b1_imports(month_range: tuple[int, int]) -> dict[str, dict[str, dict[str, float]]] | None:
    """Scrape B1 HTML page and return import data for tracked products.

    The B1 page can span several calendar months (currently year-to-date), so
    rows are filtered to ``month_range`` (inclusive) to keep imports aligned with
    the active TRQ quarter — Laura's "imports over the same time period".

    Returns:
        {
            "Hot-Rolled Sheet": {
                "China": {"tonnes": 1234.5, "value": 5678.9},
                ...
            },
            ...
        }
    """
    allowed_months = set(range(month_range[0], month_range[1] + 1))
    try:
        log.info("Downloading B1 page (%s)...", B1_URL)
        resp = SESSION.get(B1_URL, timeout=HTML_TIMEOUT)
        resp.raise_for_status()
        resp.encoding = "utf-8"
    except requests.RequestException as exc:
        log.warning("B1 download failed: %s", exc)
        return None

    try:
        return _parse_b1_page(resp.text, allowed_months)
    except Exception as exc:
        # B1 is soft-fail BY DESIGN (the sheet then says "not available"); an
        # unexpected page/format change must not take down the TRQ update.
        log.warning("B1 parsing failed (page format changed?): %s", exc)
        return None


def _parse_b1_page(html: str, allowed_months: set[int]) -> dict[str, dict[str, dict[str, float]]] | None:
    """Parse the B1 HTML into {product: {country: {tonnes, value}}}, keeping
    only rows whose Month falls in ``allowed_months``."""
    log.info("Parsing B1 HTML (%.1f MB)...", len(html) / 1_000_000)
    soup = BeautifulSoup(html, "lxml")

    # Find the main data table (the one with >100 rows)
    main_table = None
    for table in soup.find_all("table"):
        if len(table.find_all("tr")) > 100:
            main_table = table
            break

    if main_table is None:
        log.warning("Could not find main data table in B1 page")
        return None

    rows = main_table.find_all("tr")
    log.info("B1 table has %d rows", len(rows))

    # Stateful parsing
    result: dict[str, dict[str, dict[str, float]]] = {}
    current_hs: str | None = None
    current_month: str | None = None

    def process(hs_code: str, month: str, country: str, tonnes_str: str, value_str: str):
        product = _HTS_REVERSE.get(hs_code[:8])
        if product is None:
            return  # Not a tracked HS code
        try:
            if int(month) not in allowed_months:
                return  # Outside the quarter-aligned month window
        except (ValueError, TypeError):
            return  # Unparseable month — skip rather than miscount
        country = COUNTRY_NAME_MAP.get(country, country)
        tonnes = parse_number(tonnes_str)
        value = parse_number(value_str)
        if product not in result:
            result[product] = {}
        if country not in result[product]:
            result[product][country] = {"tonnes": 0.0, "value": 0.0}
        result[product][country]["tonnes"] += tonnes
        result[product][country]["value"] += value

    for row in rows:
        cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
        n = len(cells)

        if n == 8 and len(cells[1]) == 10 and cells[1].isdigit():
            current_hs = cells[1]
            current_month = cells[3]
            process(current_hs, current_month, cells[4], cells[5], cells[6])

        elif n == 6 and cells[1].isdigit() and len(cells[1]) <= 2 and 1 <= int(cells[1]) <= 12:
            current_month = cells[1]
            process(current_hs or "", current_month, cells[2], cells[3], cells[4])

        elif n == 5 and current_hs:
            process(current_hs, current_month or "", cells[1], cells[2], cells[3])

        elif n == 7 and "Summary" in cells[1]:
            current_hs = None
            continue

    log.info("B1 parsing complete: %d tracked products found", len(result))
    return result if result else None


# ---------------------------------------------------------------------------
# Excel Workbook Management
# ---------------------------------------------------------------------------

# Weekly utilization snapshots occupy column E onward (A=product, B=country,
# C=max quota, D=max share); the OVER flag sits right after the newest date.
FIRST_DATE_COL = 5


def _write_text(ws, row: int, col: int, text):
    """Write scraped text defensively: openpyxl stores any string starting
    with '=' as a formula, so an odd/hostile country name could inject a
    formula into the published workbook. Force literal string storage."""
    cell = ws.cell(row=row, column=col, value=text)
    if isinstance(text, str) and text.startswith("="):
        cell.data_type = "s"
    return cell


def _write_total_cell(ws, row: int, col: int,
                      first_row: int | None, last_row: int | None):
    """Write one product's TOTAL cell for one date column: a SUM over its
    country rows, or a literal 0.0 when there are no country rows — a SUM
    range would then include the TOTAL cell itself (an Excel circular
    reference, which is exactly what a zero-utilization product produced
    before this helper existed)."""
    cell = ws.cell(row=row, column=col)
    if first_row is not None and last_row is not None and first_row <= last_row < row:
        letter = get_column_letter(col)
        cell.value = f"=SUM({letter}{first_row}:{letter}{last_row})"
    else:
        cell.value = 0.0
    cell.number_format = "0.0%"
    cell.fill = YELLOW_FILL
    return cell


def _rewrite_total_formulas(ws, first_date_col: int, last_date_col: int):
    """Rewrite every product's TOTAL cell in every date column from the
    sheet's CURRENT physical layout.

    ws.insert_rows() shifts cell contents but never rewrites formula text, so
    any TOTAL formula written before an insertion above it keeps a stale row
    range (docs/known-issues.md Bug 1 — the original fix repaired only the
    in-memory ranges cache, not formulas already on the sheet). Country VALUES
    always sit on the correct rows, so re-deriving each SUM range from the
    final layout is a pure repair and is idempotent when nothing shifted."""
    ranges = _get_product_row_ranges(ws)
    for r in ranges.values():
        total_row = r["total_row"]
        if not total_row:
            continue
        first, last = r["first_row"], r["last_country_row"]
        if last >= total_row:  # zero-country product: its block IS the TOTAL row
            first, last = None, None
        for col in range(first_date_col, last_date_col + 1):
            _write_total_cell(ws, total_row, col, first, last)


def _find_over_col(ws) -> int | None:
    """Find the column index (1-based) of the OVER header in row 2."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and str(val).strip().upper() == "OVER":
            return col
    return None


def _find_date_col(ws, date_header: str) -> int | None:
    """Find the column index of a specific date header in row 2."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val and str(val).strip() == date_header:
            return col
    return None


def _get_product_row_ranges(ws) -> dict[str, dict]:
    """Scan the sheet to build a map of product -> {first_row, last_country_row, total_row}.

    Assumes data starts at row 3. Column A = product name, Column B = country.
    """
    ranges: dict[str, dict] = {}
    current_product = None
    first_row = None
    last_country_row = None

    for row in range(3, ws.max_row + 1):
        prod = ws.cell(row=row, column=1).value
        country = ws.cell(row=row, column=2).value

        if prod and str(prod).strip():
            prod = str(prod).strip()
            if prod != current_product:
                # Finalize previous product
                if current_product and first_row:
                    ranges[current_product]["last_country_row"] = last_country_row
                current_product = prod
                first_row = row
                last_country_row = row
                if prod not in ranges:
                    ranges[prod] = {"first_row": row, "last_country_row": row, "total_row": None}

        if country and str(country).strip() == "TOTAL" and current_product:
            ranges[current_product]["total_row"] = row
            ranges[current_product]["last_country_row"] = last_country_row
        elif country and str(country).strip() != "TOTAL" and current_product:
            last_country_row = row
            ranges[current_product]["last_country_row"] = last_country_row

    return ranges


def create_trq_sheet(wb: Workbook, sheet_name: str, quarter: str,
                     trq_type: str, data: dict, today: date,
                     start_date: str | None = None, end_date: str | None = None):
    """Create a new TRQ sheet with initial data.

    start_date/end_date are the authoritative human-readable quarter dates from
    landing-page discovery; when omitted, fall back to the (stale) computed
    range. Only the row-1 title cell uses them."""
    ws = wb.create_sheet(title=sheet_name)
    if start_date is None or end_date is None:
        start_date, end_date = get_quarter_date_range(quarter, today)
    type_label = "Non-FTA" if trq_type == "NFTA" else "FTA"

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"{type_label} Quota - Quarter {quarter[1]}: {start_date} to {end_date}")

    # Row 2: Headers
    date_header = format_date_header(today)
    headers = [None, "Country", "Max Quota (KG)", "Max Share (%)", date_header, "OVER"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        if header is not None:
            cell.font = BOLD_FONT
    # Yellow on OVER header
    ws.cell(row=2, column=6).fill = YELLOW_FILL

    # Data rows
    row_num = 3
    date_col = FIRST_DATE_COL  # Column E
    over_col = FIRST_DATE_COL + 1  # Column F

    for product in TRACKED_PRODUCTS:
        prod_data = data.get(product, {
            "max_quota": 0, "max_share": 0.0, "countries": {},
        })
        countries = prod_data.get("countries", {})
        max_quota = prod_data.get("max_quota", 0)
        max_share = prod_data.get("max_share", 0.0)

        # Sort countries alphabetically for consistent ordering
        sorted_countries = sorted(countries.keys())

        if not sorted_countries:
            # No countries — just write TOTAL row with 0
            ws.cell(row=row_num, column=1, value=product)
            ws.cell(row=row_num, column=2, value="TOTAL")
            _write_total_cell(ws, row_num, date_col, None, None)
            row_num += 1
            continue

        first_country_row = row_num
        for country in sorted_countries:
            share = countries[country]
            ws.cell(row=row_num, column=1, value=product)
            _write_text(ws, row_num, 2, country)

            quota_cell = ws.cell(row=row_num, column=3, value=max_quota)
            quota_cell.number_format = "#,##0"

            share_cap_cell = ws.cell(row=row_num, column=4, value=max_share)
            share_cap_cell.number_format = "0.0%"

            util_cell = ws.cell(row=row_num, column=date_col, value=share)
            util_cell.number_format = "0.0%"

            # OVER flag
            if share >= max_share and max_share > 0:
                ws.cell(row=row_num, column=over_col, value="YES")

            row_num += 1

        last_country_row = row_num - 1

        # TOTAL row
        ws.cell(row=row_num, column=1, value=product)
        ws.cell(row=row_num, column=2, value="TOTAL")
        _write_total_cell(ws, row_num, date_col, first_country_row, last_country_row)
        row_num += 1

    log.info("Created sheet '%s' with %d rows", sheet_name, row_num - 1)


def update_trq_sheet(ws, data: dict, today: date):
    """Add a new date column to an existing TRQ sheet."""
    date_header = format_date_header(today)

    # Check if this date already exists
    if _find_date_col(ws, date_header):
        log.info("Sheet '%s' already has column '%s' — skipping", ws.title, date_header)
        return

    # Find OVER column
    over_col = _find_over_col(ws)
    if over_col is None:
        log.warning("No OVER column found in '%s' — cannot update", ws.title)
        return

    # New date column replaces OVER position; OVER moves right by 1
    new_date_col = over_col
    new_over_col = over_col + 1

    # Write date header
    header_cell = ws.cell(row=2, column=new_date_col, value=date_header)
    header_cell.font = BOLD_FONT

    # Clear the entire new date column first (old OVER "YES" values may linger)
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=new_date_col).value = None
        ws.cell(row=row, column=new_date_col).fill = PatternFill()

    # Get product row ranges
    ranges = _get_product_row_ranges(ws)

    # Write country values for the new column. TOTAL cells are NOT written
    # here: rows may still be inserted below (new countries), and insert_rows
    # shifts cells without rewriting formula text, so any formula written now
    # could go stale. All TOTAL cells are (re)written in one pass at the end,
    # from the final physical layout.
    for row in range(3, ws.max_row + 1):
        prod = ws.cell(row=row, column=1).value
        country = ws.cell(row=row, column=2).value

        if not prod or not country:
            continue

        prod = str(prod).strip()
        country = str(country).strip()

        if country != "TOTAL":
            # Country row — find share value
            prod_data = data.get(prod, {})
            countries = prod_data.get("countries", {})
            share = countries.get(country)

            if share is not None:
                cell = ws.cell(row=row, column=new_date_col)
                cell.value = share
                cell.number_format = "0.0%"
            # If country not in current data, leave blank

    # Check for new countries not yet in the sheet
    for prod in TRACKED_PRODUCTS:
        prod_data = data.get(prod, {})
        countries = prod_data.get("countries", {})

        if prod not in ranges:
            continue

        existing_countries = set()
        r = ranges[prod]
        for row in range(r["first_row"], (r["total_row"] or r["last_country_row"]) + 1):
            c = ws.cell(row=row, column=2).value
            if c and str(c).strip() != "TOTAL":
                existing_countries.add(str(c).strip())

        for country in sorted(countries.keys()):
            if country not in existing_countries:
                total_row = r["total_row"]
                if total_row:
                    log.info("New country '%s' for product '%s' — appending row", country, prod)
                    ws.insert_rows(total_row)
                    new_row = total_row

                    # Update current product's ranges
                    r["total_row"] = total_row + 1
                    r["last_country_row"] = new_row

                    # Shift ALL subsequent products' row references down by 1
                    for other_prod, other_r in ranges.items():
                        if other_prod == prod:
                            continue
                        if other_r["first_row"] > total_row:
                            other_r["first_row"] += 1
                        if other_r["last_country_row"] and other_r["last_country_row"] >= total_row:
                            other_r["last_country_row"] += 1
                        if other_r["total_row"] and other_r["total_row"] > total_row:
                            other_r["total_row"] += 1

                    ws.cell(row=new_row, column=1, value=prod)
                    _write_text(ws, new_row, 2, country)

                    quota_cell = ws.cell(row=new_row, column=3, value=prod_data.get("max_quota", 0))
                    quota_cell.number_format = "#,##0"

                    share_cap_cell = ws.cell(row=new_row, column=4, value=prod_data.get("max_share", 0.0))
                    share_cap_cell.number_format = "0.0%"

                    cell = ws.cell(row=new_row, column=new_date_col, value=countries[country])
                    cell.number_format = "0.0%"
                else:
                    log.warning("New country '%s' for product '%s' has no TOTAL row "
                                "to insert above — row NOT added", country, prod)

    # All rows are in their final positions now. Rewrite every TOTAL cell in
    # every date column from the final layout — this writes the new column's
    # TOTALs for the first time, repairs any formula made stale by the inserts
    # above (in the new AND all historical columns), and heals corruption left
    # by earlier runs of the pre-fix code.
    _rewrite_total_formulas(ws, FIRST_DATE_COL, new_date_col)

    # Rewrite OVER column in new position (one column to the right)
    # The old OVER column is now the new date column — do NOT clear it.
    over_header = ws.cell(row=2, column=new_over_col, value="OVER")
    over_header.font = BOLD_FONT
    over_header.fill = YELLOW_FILL

    # Recalculate OVER for all data rows
    for row in range(3, ws.max_row + 1):
        country = ws.cell(row=row, column=2).value
        if not country or str(country).strip() == "TOTAL":
            ws.cell(row=row, column=new_over_col).value = None
            continue

        max_share = ws.cell(row=row, column=4).value
        if max_share is None or max_share == 0:
            ws.cell(row=row, column=new_over_col).value = None
            continue

        latest_share = ws.cell(row=row, column=new_date_col).value
        if latest_share is not None and isinstance(latest_share, (int, float)):
            ws.cell(row=row, column=new_over_col, value="YES" if latest_share >= max_share else None)
        else:
            ws.cell(row=row, column=new_over_col).value = None

    log.info("Updated sheet '%s' with column '%s'", ws.title, date_header)


def create_b1_sheet(wb: Workbook, b1_data: dict | None, trq_quarter: str, today: date,
                    quarter_end: date | None = None):
    """Create or rewrite the B1 Imports sheet.

    quarter_end is the discovered TRQ quarter's end date; the B1 calendar
    window always falls in that date's year (true for all four quarters,
    including year-wrapping Q3), so it labels the window correctly even when
    today lies outside the window (early Q3 in December, or a stale landing
    page leaving last quarter selected past its end)."""
    sheet_name = "B1 Imports"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # Determine calendar quarter label. Prefer the quarter's real end date
    # (window year == end-date year, all quarters); fall back to a heuristic
    # on today only if no end date was provided.
    cal_start_m, cal_end_m = TRQ_TO_B1_CALENDAR_MONTHS[trq_quarter]
    if quarter_end is not None:
        year = quarter_end.year
    else:
        year = today.year + 1 if today.month > cal_end_m else today.year
    cal_label = (f"{_MONTH_NAMES[cal_start_m]} 1 to {_MONTH_NAMES[cal_end_m]} "
                 f"{30 if cal_end_m in (4, 6, 9, 11) else 31}, {year}")

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"B1 Import Data - Calendar Quarter: {cal_label}")

    # Rows 3-8: Disclaimers
    ws.cell(row=3, column=1,
            value="NOTE: B1 data reflects actual customs entries using calendar quarters (Jan-Mar, Apr-Jun, etc.).")
    ws.cell(row=4, column=1,
            value="TRQ utilization data uses offset quarters (Mar 26-Jun 26, etc.) and tracks permit issuance, not customs entries.")
    ws.cell(row=5, column=1,
            value="The two datasets have ~5-6 day boundary misalignment and measure different aspects of the import process.")
    ws.cell(row=6, column=1,
            value="B1 imports are matched at the 8-digit tariff item level using the HTS codes listed in the 'HTS code covered' sheet.")
    ws.cell(row=7, column=1,
            value="These HTS codes are a selected subset; the official TRQ product scope may include additional tariff items not tracked here.")
    ws.cell(row=8, column=1,
            value=f"Imports are limited to calendar months {_MONTH_NAMES[cal_start_m]}-{_MONTH_NAMES[cal_end_m]} to align with the TRQ quarter; the quarter may be partially complete (data current through the latest available customs month).")

    # Row 9: Headers
    headers = ["Product Category", "Country", "Total Tonnes", "Total C$1000", "Avg C$/Tonne"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = BOLD_FONT

    if b1_data is None:
        ws.cell(row=11, column=1, value="B1 data not available for the matching calendar quarter.")
        log.info("B1 sheet created (no data)")
        return

    # Write data
    row_num = 10
    for product in TRACKED_PRODUCTS:
        if product not in b1_data:
            continue
        countries = b1_data[product]
        for country in sorted(countries.keys()):
            d = countries[country]
            tonnes = d["tonnes"]
            value = d["value"]
            avg_price = value * 1000 / tonnes if tonnes > 0 else 0.0

            ws.cell(row=row_num, column=1, value=product)
            _write_text(ws, row_num, 2, country)

            t_cell = ws.cell(row=row_num, column=3, value=tonnes)
            t_cell.number_format = "#,##0.00"

            v_cell = ws.cell(row=row_num, column=4, value=value)
            v_cell.number_format = "#,##0.00"

            p_cell = ws.cell(row=row_num, column=5, value=avg_price)
            p_cell.number_format = "#,##0.00"

            row_num += 1

    log.info("B1 sheet created with %d data rows", row_num - 10)


def create_hts_sheet(wb: Workbook):
    """Create the static HTS code reference sheet."""
    sheet_name = "HTS code covered"
    if sheet_name in wb.sheetnames:
        return  # Already exists

    ws = wb.create_sheet(title=sheet_name)

    # Flat products (column A) and Long products (column B)
    flat_products = ["Hot-Rolled Sheet", "Steel Plate", "Cold-Rolled Sheet", "Coated Steel Sheet"]
    long_products = ["Rebar", "Hot-Rolled Bar", "Wire Rod", "Structural Steel"]

    row = 1
    # Write flat products
    for product in flat_products:
        ws.cell(row=row, column=1, value=product)
        row += 1
        for code in HTS_CODE_MAP[product]:
            ws.cell(row=row, column=1, value=code)
            row += 1
        row += 1  # blank row between products

    # Write long products in column B (restart from row 1)
    row = 1
    for product in long_products:
        ws.cell(row=row, column=2, value=product)
        row += 1
        for code in HTS_CODE_MAP[product]:
            ws.cell(row=row, column=2, value=code)
            row += 1
        row += 1

    log.info("HTS code reference sheet created")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    today = date.today()

    # Discover the CURRENT-quarter reports from the official landing page. The
    # source renames quarter files across program years (e.g. Q1 -> Y2Q1) and
    # shifts boundary dates, so the filename, label, and dates are READ from the
    # page, never constructed. FAIL LOUD on any problem -- never silently fall
    # back to a stale/closed quarter, which is exactly what hid the rollover bug.
    try:
        reports = discover_current_reports(today)
    except RuntimeError as exc:
        log.error("TRQ report discovery failed: %s", exc)
        sys.exit(1)

    quarter = reports["FTA"]["quarter"]
    log.info("Today: %s, current TRQ quarter: %s (%s to %s)",
             today, quarter, reports["FTA"]["start_str"], reports["FTA"]["end_str"])

    # Download the discovered CSVs (full URLs taken verbatim from the page).
    fta_csv = download_csv(reports["FTA"]["csv_url"])
    nfta_csv = download_csv(reports["NFTA"]["csv_url"])
    if fta_csv is None or nfta_csv is None:
        log.error("Could not download current TRQ CSV data (FTA=%s, NFTA=%s). Exiting.",
                  reports["FTA"]["csv_url"], reports["NFTA"]["csv_url"])
        sys.exit(1)

    # Parse TRQ data. FAIL LOUD on any parse problem — a structurally changed
    # CSV must never be written to the workbook as plausible-looking data.
    try:
        log.info("Parsing FTA CSV...")
        fta_data = parse_trq_csv(fta_csv)
        log.info("Parsing NFTA CSV...")
        nfta_data = parse_trq_csv(nfta_csv)
    except Exception as exc:
        log.error("TRQ CSV parse failed: %s", exc)
        sys.exit(1)

    # B1 import data. The TRQ "Q1..Q4" label maps to calendar months in a
    # year-independent way (TRQ_TO_B1_CALENDAR_MONTHS), so this is unchanged.
    b1_data = None
    if should_fetch_b1(quarter, today):
        b1_data = scrape_b1_imports(TRQ_TO_B1_CALENDAR_MONTHS[quarter])
        if b1_data is None:
            log.warning("B1 scraping returned no data for tracked products")
    else:
        log.info("B1 data for matching calendar quarter not yet available (TRQ %s, current month %d)",
                 quarter, today.month)

    # Load or create workbook
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(OUTPUT_FILE):
        log.info("Loading existing workbook: %s", OUTPUT_FILE)
        try:
            wb = load_workbook(OUTPUT_FILE)
        except Exception as exc:
            log.error("Could not open existing workbook %s: %s", OUTPUT_FILE, exc)
            sys.exit(1)
    else:
        log.info("Creating new workbook")
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # On a quarter transition, preserve the previous quarter's B1 imports. The
    # single "B1 Imports" sheet is deleted & rewritten every run, so without this
    # the just-closed quarter's customs data would be silently overwritten with
    # "not available" until the new quarter's B1 calendar window opens.
    trq_sheets = [f"{TRQ_SHEET_PREFIX['NFTA']} {quarter}", f"{TRQ_SHEET_PREFIX['FTA']} {quarter}"]
    if not any(s in wb.sheetnames for s in trq_sheets) and "B1 Imports" in wb.sheetnames:
        prev_q = {"Q1": "Q4", "Q2": "Q1", "Q3": "Q2", "Q4": "Q3"}[quarter]
        archive = f"B1 Imports {prev_q}"
        if archive in wb.sheetnames:
            del wb[archive]
        wb["B1 Imports"].title = archive
        log.info("Quarter transition: archived previous B1 data as '%s'", archive)

    # Update or create TRQ sheets
    for trq_type, data in [("NFTA", nfta_data), ("FTA", fta_data)]:
        sheet_name = f"{TRQ_SHEET_PREFIX[trq_type]} {quarter}"
        rep = reports[trq_type]
        if sheet_name in wb.sheetnames:
            log.info("Updating existing sheet '%s'...", sheet_name)
            update_trq_sheet(wb[sheet_name], data, today)
        else:
            log.info("Creating new sheet '%s'...", sheet_name)
            create_trq_sheet(wb, sheet_name, quarter, trq_type, data, today,
                             start_date=rep["start_str"], end_date=rep["end_str"])

    # B1 sheet
    create_b1_sheet(wb, b1_data, quarter, today, quarter_end=reports["FTA"]["end_date"])

    # HTS reference sheet
    create_hts_sheet(wb)

    # Save atomically: write to a temp file, then swap it in. A crash mid-save
    # must never truncate the accumulated history file. On local Windows runs
    # the destination can be transiently locked (Excel has it open, or the
    # OneDrive sync engine holds a handle) — retry briefly, and on final
    # failure keep the .tmp (it holds the new data; *.tmp is gitignored).
    tmp_file = OUTPUT_FILE + ".tmp"
    wb.save(tmp_file)
    for attempt in range(3):
        try:
            os.replace(tmp_file, OUTPUT_FILE)
            break
        except PermissionError as exc:
            if attempt == 2:
                log.error("Could not replace %s (locked by Excel/OneDrive?): %s — "
                          "new data preserved at %s", OUTPUT_FILE, exc, tmp_file)
                sys.exit(1)
            time.sleep(2)
    log.info("Workbook saved to %s", OUTPUT_FILE)
    log.info("Done.")


if __name__ == "__main__":
    main()
