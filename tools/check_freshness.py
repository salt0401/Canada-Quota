# -*- coding: utf-8 -*-
"""
Report the newest weekly date column in the TRQ workbook, and optionally assert
it is fresh enough.

Two callers, one fact:

  * tools/server-weekly-task.ps1 runs it with --expect-date after the scrape, to
    prove the run actually advanced the workbook rather than re-saving an
    unchanged one. "The script exited 0" and "a new column exists" are different
    claims -- the whole known-issues.md file is about that gap.
  * .github/workflows/data-freshness-watchdog.yml runs it with --max-age-days as
    the external heartbeat, because nothing on the company server notices a
    scheduled task that never fires.

DELIBERATELY STANDALONE: it imports openpyxl and nothing else from this project.
The watchdog runs on a GitHub runner where installing requests/bs4/lxml just to
read two cells would be waste, and importing canada_trq_tracker would build an
HTTP session as a module-level side effect. The two layout constants below are
duplicated from that module for the same reason; they are part of the workbook's
on-disk format, which is fixed by the shipped file, not by this script.

Usage:
    python tools/check_freshness.py
    python tools/check_freshness.py --expect-date 2026-08-03
    python tools/check_freshness.py --max-age-days 8
    python tools/check_freshness.py --json
"""

import argparse
import json
import os
import re
import sys
from datetime import date, datetime, timezone

from openpyxl import load_workbook

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_WORKBOOK = os.path.join(_PROJECT_ROOT, "data", "canada_trq_tracker.xlsx")

# Mirrors canada_trq_tracker.py: TRQ sheets are named "<prefix> <quarter>", and
# the dated columns are headed in row 2. See the module docstring on why these
# are duplicated rather than imported.
TRQ_SHEET_PREFIXES = ("non-FTA ", "FTA ")
HEADER_ROW = 2

_MONTHS = {
    "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
    "July": 7, "August": 8, "September": 9, "October": 10, "November": 11,
    "December": 12,
}
# format_date_header() writes "July 27 2026" -- full month, no leading zero, no
# comma. Anchored so a stray header cell cannot half-match.
_DATE_HEADER_RE = re.compile(r"^([A-Z][a-z]+) (\d{1,2}) (\d{4})$")


def parse_date_header(value) -> date | None:
    """Turn a 'July 27 2026' header back into a date, or None if it is not one."""
    if value is None:
        return None
    m = _DATE_HEADER_RE.match(str(value).strip())
    if not m:
        return None
    month = _MONTHS.get(m.group(1))
    if month is None:
        return None
    try:
        return date(int(m.group(3)), month, int(m.group(2)))
    except ValueError:
        return None


def collect_date_columns(path: str) -> dict:
    """Map each TRQ sheet to the sorted list of dates it carries columns for."""
    if not os.path.exists(path):
        raise RuntimeError(f"Workbook not found: {path}")
    # read_only + data_only: we want header strings, never formula evaluation,
    # and the file can be a few MB once a quarter's history accumulates.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        found = {}
        for name in wb.sheetnames:
            if not name.startswith(TRQ_SHEET_PREFIXES):
                continue
            ws = wb[name]
            dates = []
            for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW):
                for cell in row:
                    d = parse_date_header(cell.value)
                    if d is not None:
                        dates.append(d)
            found[name] = sorted(set(dates))
        return found
    finally:
        wb.close()


def newest_date(path: str) -> tuple[date | None, dict]:
    per_sheet = collect_date_columns(path)
    everything = [d for dates in per_sheet.values() for d in dates]
    return (max(everything) if everything else None), per_sheet


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Report/assert the newest weekly column in the TRQ workbook.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--expect-date", metavar="YYYY-MM-DD",
                        help="fail unless the newest column is exactly this date")
    parser.add_argument("--max-age-days", type=int, metavar="N",
                        help="fail if the newest column is more than N days old "
                             "(compared against today in UTC)")
    parser.add_argument("--json", action="store_true",
                        help="emit a machine-readable summary")
    args = parser.parse_args(argv)

    try:
        newest, per_sheet = newest_date(args.workbook)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    today = datetime.now(timezone.utc).date()
    age = (today - newest).days if newest else None

    if args.json:
        print(json.dumps({
            "workbook": args.workbook,
            "newest_column": newest.isoformat() if newest else None,
            "today_utc": today.isoformat(),
            "age_days": age,
            "sheets": {k: [d.isoformat() for d in v] for k, v in per_sheet.items()},
        }, indent=2))
    else:
        for name, dates in per_sheet.items():
            span = f"{dates[0]} .. {dates[-1]}" if dates else "(no dated columns)"
            print(f"  {name:<20} {len(dates):>3} column(s)   {span}")
        print(f"  newest column: {newest}   today (UTC): {today}   age: {age} day(s)")

    if newest is None:
        print("FAIL: the workbook has no dated columns at all.", file=sys.stderr)
        return 1

    if args.expect_date:
        want = date.fromisoformat(args.expect_date)
        if newest != want:
            print(f"FAIL: newest column is {newest}, expected {want}. The run did "
                  f"not add today's column.", file=sys.stderr)
            return 1

    if args.max_age_days is not None and age > args.max_age_days:
        print(f"FAIL: newest column {newest} is {age} days old "
              f"(limit {args.max_age_days}).", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
